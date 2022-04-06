import matplotlib.pyplot as plt
import openpyxl
from sklearn_extra.cluster import KMedoids
from sklearn.cluster import OPTICS
import numpy as np
import pandas as pd

from utils.dataset_parsing import simulations_dataset as ds
from utils import scatter_plot
from feature_extraction.weighted_pca import clustering as cl
from sklearn.cluster import Birch
from sklearn.decomposition import pca
from sklearn.metrics.cluster import adjusted_rand_score
from sklearn.metrics.cluster import adjusted_mutual_info_score
from sklearn.metrics.cluster import fowlkes_mallows_score
from sklearn.metrics.cluster import v_measure_score
from sklearn.metrics.cluster import homogeneity_score

from numpy import unique
from numpy import where
from matplotlib import pyplot
from sklearn.datasets import make_classification
from sklearn.mixture import GaussianMixture

from numpy import unique
from numpy import where
from matplotlib import pyplot
from sklearn.datasets import make_classification
from sklearn.cluster import AffinityPropagation

from sklearn.cluster import MeanShift

from sklearn.cluster import AgglomerativeClustering

import seaborn as sns
from sklearn.datasets.samples_generator import make_blobs

from utils.benchmark import clustering_evaluation_richard_metric as cvrm
import collections
import statistics


# Previous work
# PCA 2d
# spikes, labels = ds.get_dataset_simulation_pca_2d(simNr=1)

# PCA 3d
# spikes, labels = ds.get_dataset_simulation_pca_3d(simNr=1)

# labels_knn = cl.apply_clustering_algorithm(spikes,labels,0)
# labels_SBM = cl.apply_clustering_algorithm(spikes,labels,1)

# scatter_plot.plot('Ground Truth', spikes, labels, marker='o')
# scatter_plot.plot('KNN', spikes, labels_knn, marker='o')
# scatter_plot.plot('SBM', spikes, labels_SBM, marker='o')


# MIHM : algoritmi de clusterizare
# kmedoids = KMedoids(n_clusters=int(np.amax(labels)) + 1, random_state=0).fit(spikes)
# labels_kmedoids = kmedoids.labels_

# birch = Birch(branching_factor = 50, n_clusters = int(np.amax(labels)) + 1, threshold = 0.75).fit(spikes)
# labels_birch = birch.labels_

# gaussian_model = GaussianMixture(n_components=int(np.amax(labels)) + 1).fit(spikes)
# labels_gaussian = gaussian_model.predict(spikes)

# affinity_model = AffinityPropagation(damping=0.98).fit(spikes)
# labels_affinity_propagation = affinity_model.labels_

# mean_shift_model = MeanShift().fit(spikes)
# labels_mean_shift = mean_shift_model.labels_

# agglomerative_model = AgglomerativeClustering(n_clusters=int(np.amax(labels)) + 1).fit(spikes)
# labels_agglomerative = agglomerative_model.labels_

# scatter_plot.plot('Meoids', spikes, labels_kmedoids, marker='o')
# scatter_plot.plot('Birch', spikes, labels_birch, marker='o')
# scatter_plot.plot('Gaussian', spikes, labels_gaussian, marker='o')
# scatter_plot.plot('Affinity Propaggation', spikes, labels_affinity_propagation, marker='o')
# scatter_plot.plot('Mean Shift', spikes, labels_mean_shift, marker='o')
# scatter_plot.plot('Agglomerative', spikes, labels_agglomerative, marker='o')
# scatter_plot.plot('Optics', spikes, labels_optics, marker='o')
# scatter_plot.plot('Optics', noise, labels_optics2, marker='o')

# plt.show()


def gaussian_mixture_clustering(underclustered_cluster, bic, aic):
    """
    Recluster the chosen cluster by performing gaussian clustering twice, once using the bic value and once using the
    aic value as the parameter representing the number of subclusters to be found

    :param underclustered_cluster: The cluster to be reclustered using gaussian clustering and the bic and aic values
    :param bic: The "bic" value of the cluster
    :param aic: The "aic" value of the cluster
    :return: The two sets of labels obtained from the two clusterings
    """
    gmm = GaussianMixture(n_components=bic)
    gmm.fit(underclustered_cluster)
    labels_gaussian_bic = gmm.predict(underclustered_cluster)
    scatter_plot.plot('Gaussian', underclustered_cluster, labels_gaussian_bic, marker='o')

    gmm = GaussianMixture(n_components=aic)
    gmm.fit(underclustered_cluster)
    labels_gaussian_aic = gmm.predict(underclustered_cluster)
    scatter_plot.plot('Gaussian', underclustered_cluster, labels_gaussian_aic, marker='o')

    return labels_gaussian_bic, labels_gaussian_aic


def gaussian_mixture_bic(cluster):
    """
    Calculate the "bic" value of a cluster

    :param cluster: The cluster whose "bic" value is calculated
    :return: The bic value of the cluster
    """
    n_components = np.arange(1, 21)
    models = [GaussianMixture(n, covariance_type='full', random_state=0).fit(cluster) for n in
              n_components]
    pos_min = 1
    pos_cur = 1
    minim = models[0].bic(cluster)
    for m in models:
        if minim > m.bic(cluster):
            minim = m.bic(cluster)
            pos_min = pos_cur
        pos_cur += 1
    print(pos_min)
    bic = pos_min
    return bic


def gaussian_mixture_aic(cluster):
    """
    Calculate the "aic" value of a cluster

    :param cluster: The cluster whose "aic" value is calculated
    :return: The aic value of the cluster
    """
    n_components = np.arange(1, 21)
    models = [GaussianMixture(n, covariance_type='full', random_state=0).fit(cluster) for n in
              n_components]
    pos_min = 1
    pos_cur = 1
    minim = models[0].aic(cluster)
    for m in models:
        if minim > m.aic(cluster):
            minim = m.aic(cluster)
            pos_min = pos_cur
        pos_cur += 1
    print(pos_min)
    aic = pos_min
    return aic


def get_dataset_simulation_pca_x_dimensions(sim_nr, spike_length=79, align_to_peak=True, normalize_spike=False,
                                            nr_of_dimensions_x=2):
    """
    Get the spikes and corresponding labels of the desired simulation reduced by using PCA to the desired number of
    dimensions

    :param sim_nr: The number of the simulation
    :param spike_length: The length (number of dimensions) of the spikes
    :param align_to_peak: The parameter that decides whether the spikes are aligned to peak
    :param normalize_spike: The parameter that decides whether the spikes are being normalized
    :param nr_of_dimensions_x: The number of dimensions of the spikes returned by the function
    :return: Returns the spikes in the chosen number of dimensions and their corresponding labels
    """
    spikes, labels = ds.get_dataset_simulation(sim_nr, spike_length, align_to_peak, normalize_spike)
    pca_xd = pca.PCA(n_components=nr_of_dimensions_x)
    spikes_pca_xd = pca_xd.fit_transform(spikes)
    return spikes_pca_xd, labels


def update_excel_cell(searched_value_row, searched_value_col, path, sheet, new_value):
    """
    Update a microsoft excel cell that is found at the intersection of two cells containing certain values inside an
    excel file found at a certain path

    :param searched_value_row: The value inside a cell that the row of the cell that is to be updated has to contain
    :param searched_value_col: The value inside a cell that the column of the cell that is to be updated has to contain
    :param path: The path where the excel file is found
    :param sheet: The sheet in the excel file where the cell that is to be updated is found
    :param new_value: The value that will be written inside the cell that is to be updated
    :return:
    """
    wb_obj = openpyxl.load_workbook(path.strip())
    sh_obj = wb_obj[sheet]
    searched_row = -1
    searched_col = -1
    if searched_value_col:
        for row in sh_obj.iter_rows(min_row=sh_obj.min_row, max_row=sh_obj.max_row + 1, min_col=sh_obj.min_column,
                                    max_col=sh_obj.max_column + 1):
            for cell in row:
                if cell.value == searched_value_col:
                    searched_col = cell.column
                    break
            if searched_col != -1:
                break
    if searched_value_row:
        for col in sh_obj.iter_cols(min_row=sh_obj.min_row, max_row=sh_obj.max_row + 1, min_col=sh_obj.min_column,
                                    max_col=sh_obj.max_column + 1):
            for cell in col:
                if cell.value == searched_value_row:
                    searched_row = cell.row
                    break
            if searched_row != -1:
                break
    if searched_row != -1 and searched_col != -1:
        sh_obj.cell(searched_row, searched_col).value = new_value
        wb_obj.save(path)
    elif searched_col != -1:
        for row in sh_obj.iter_rows(min_row=sh_obj.min_row, max_row=sh_obj.max_row + 1, min_col=searched_col,
                                    max_col=searched_col):
            cell = row[0]
            if cell.value is None:
                cell.value = new_value
                break
        wb_obj.save(path)


def write_metrics_excel(nr_of_dim, sim, nr_of_decimals=3, ari=None, ami=None, fmi=None, vm=None, ss=None,
                        path='C:/Users/Cosmy/Desktop/Proiect Master/test.xlsx'):
    """
    Write the values of up to five different clustering metrics inside a specific table found in an excel file that
    classifies the values based on the metrics type as well as the number of the simulation and the number of dimensions
    that correspond to the data from which the results emerged.

    :param nr_of_dim: The number of dimensions of the data from which the values of the metrics were obtained
    :param sim: The number of the simulation that the data from which the values of the metrics were obtained belongs to
    :param nr_of_decimals: The parameter responsible for the number of decimals that the written result will have
    :param ari: The value of the "ari" metric
    :param ami: The value of the "ami" metric
    :param fmi: The value of the "fmi" metric
    :param vm: The value of the "vm" metric
    :param ss: The values of the "ss" metric
    :param path: The path where the excel file is found
    :return:
    """
    # print("\nMetrics for underclustered cluster:")
    if ari:
        # print("Adjusted Rand Index (ARI): ", ari)
        update_excel_cell(('PCA' + str(nr_of_dim)), 'ARI', path, ('Sim' + str(sim)), round(ari, nr_of_decimals))

    if ami:
        # print("Adjusted Mutual Info Score (AMI): ", ami)
        update_excel_cell(('PCA' + str(nr_of_dim)), 'AMI', path, ('Sim' + str(sim)), round(ami, nr_of_decimals))

    if fmi:
        # print("Fowlkes Mallows Index (FMI): ", fmi)
        update_excel_cell(('PCA' + str(nr_of_dim)), 'FMI', path, ('Sim' + str(sim)), round(fmi, nr_of_decimals))

    if vm:
        # print("V-Measure (VM): ", vm)
        update_excel_cell(('PCA' + str(nr_of_dim)), 'VM', path, ('Sim' + str(sim)), round(vm, nr_of_decimals))

    if ss:
        # print("SS-Metric: ", ss)
        update_excel_cell(('PCA' + str(nr_of_dim)), 'SS', path, ('Sim' + str(sim)), round(ss, nr_of_decimals))


def write_intermediate_metrics_excel(nr_of_dim, sim, nr_of_decimals=3, ari=None, ami=None, fmi=None, vm=None, ss=None,
                                     path='C:/Users/Cosmy/Desktop/Proiect Master/test.xlsx'):
    """
    Write the values of up to five different clustering metrics inside specific columns found in an excel file that
    classifies the values based on the metrics type as well as the number of the simulation and the number of dimensions
    that correspond to the data from which the results emerged.

    :param nr_of_dim: The number of dimensions of the data from which the values of the metrics were obtained
    :param sim: The number of the simulation that the data from which the values of the metrics were obtained belongs to
    :param nr_of_decimals: The parameter responsible for the number of decimals that the written result will have
    :param ari: The value of the "ari" metric
    :param ami: The value of the "ami" metric
    :param fmi: The value of the "fmi" metric
    :param vm: The value of the "vm" metric
    :param ss: The values of the "ss" metric
    :param path: The path where the excel file is found
    :return:
    """
    # print("\nMetrics for underclustered cluster:")
    if ari:
        # print("Adjusted Rand Index (ARI): ", ari)
        update_excel_cell(None, ('ARI' + str(nr_of_dim) + 'D'), path, ('Sim' + str(sim)), round(ari, nr_of_decimals))

    if ami:
        # print("Adjusted Mutual Info Score (AMI): ", ami)
        update_excel_cell(None, ('AMI' + str(nr_of_dim) + 'D'), path, ('Sim' + str(sim)), round(ami, nr_of_decimals))

    if fmi:
        # print("Fowlkes Mallows Index (FMI): ", fmi)
        update_excel_cell(None, ('FMI' + str(nr_of_dim) + 'D'), path, ('Sim' + str(sim)), round(fmi, nr_of_decimals))

    if vm:
        # print("V-Measure (VM): ", vm)
        update_excel_cell(None, ('VM' + str(nr_of_dim) + 'D'), path, ('Sim' + str(sim)), round(vm, nr_of_decimals))

    if ss:
        # print("SS-Metric: ", ss)
        update_excel_cell(None, ('SS' + str(nr_of_dim) + 'D'), path, ('Sim' + str(sim)), round(ss, nr_of_decimals))


def calculate_metrics(labels_gt, labels_obtained, print_metrics_in_console=True,
                      ari_metric=True, ami_metric=True, fmi_metric=True, vm_metric=True, ss_metric=True):
    """
    Calculate up to five different clustering metrics

    :param labels_gt: The ground truth labels of the clusters
    :param labels_obtained: The cluster labels obtained as a result of the clustering process
    :param print_metrics_in_console: The parameter that decides whether the results will also be printed in console
    :param ari_metric: The parameter that decides whether the "ari" metric will be calculated
    :param ami_metric: The parameter that decides whether the "ami" metric will be calculated
    :param fmi_metric: The parameter that decides whether the "fmi" metric will be calculated
    :param vm_metric: The parameter that decides whether the "vm" metric will be calculated
    :param ss_metric: The parameter that decides whether the "ss" metric will be calculated
    :return: Returns the obtained values for the metrics that were calculated and "None" for those that were not
    """
    ari, ami, fmi, vm, ss = None, None, None, None, None
    if ari_metric:
        ari = adjusted_rand_score(labels_gt, labels_obtained)

    if ami_metric:
        ami = adjusted_mutual_info_score(labels_gt, labels_obtained)

    if fmi_metric:
        fmi = fowlkes_mallows_score(labels_gt, labels_obtained)

    if vm_metric:
        vm = v_measure_score(labels_gt, labels_obtained)

    if ss_metric:
        ss = cvrm.ss_metric(labels_gt, labels_obtained)

    if print_metrics_in_console:
        print("Adjusted Rand Index (ARI): ", round(ari, 3))
        print("Adjusted Mutual Info Score (AMI): ", round(ami, 3))
        print("Fowlkes Mallows Index (FMI): ", round(fmi, 3))
        print("V-Measure (VM): ", round(vm, 3))
        print("SS-Metric: ", round(ss, 3))

    return ari, ami, fmi, vm, ss


def average_metrics(spikes79_pca, labels_gt, spikes_pca_2d=None, average=20, avg_over_med=True, nr_of_subclusters=1,
                    write_intermediate_results=None, ari_metric=True, ami_metric=True, fmi_metric=True, vm_metric=True,
                    ss_metric=True):
    """
    Average or get the median value for the values obtained during multiple clustering processes for up to five
    different metrics

    :param spikes79_pca: The spikes to be labeled intro clusters
    :param labels_gt: The ground truth labels of the clusters
    :param spikes_pca_2d: The spikes to be labeled into clusters, reduced to two dimensions using PCA
    :param average: The number of gaussian clustering processes to be executed
    :param avg_over_med: The parameter that decides whether the returned results represent the averages of the results
    obtained during the clustering processes (True) or the median values of these results (False)
    :param nr_of_subclusters: The number of different clusters to be found by the gaussian clustering process
    :param write_intermediate_results: The parameter that decides whether the intermediate results (meaning the results
    found at the end of each successful clustering process) will be written in an excel file or discarded
    :param ari_metric: The parameter that decides whether the average/median of the "ari" metric will be calculated
    :param ami_metric: The parameter that decides whether the average/median of the "ami" metric will be calculated
    :param fmi_metric: The parameter that decides whether the average/median of the "fmi" metric will be calculated
    :param vm_metric: The parameter that decides whether the average/median of the "vm" metric will be calculated
    :param ss_metric: The parameter that decides whether the average/median of the "ss" metric will be calculated
    :return: Returns the obtained averaged/median values for the metrics for which one of those values were calculated
    and "0" for the metrics for which they were not
    """
    ari_avg, ami_avg, fmi_avg, vm_avg, ss_avg = 0, 0, 0, 0, 0
    metrics_values_list: list = [[0], [0], [0], [0], [0]]
    for i in range(average):
        gmm = GaussianMixture(n_components=nr_of_subclusters)
        gmm.fit(spikes79_pca)
        labels_gaussian = gmm.predict(spikes79_pca)

        # scatter_plot.plot('Gaussian', spikes20_pca_2d, labels_gaussian, marker='o')
        # scatter_plot.plot('New GT', spikes20_pca_2d, current_groundtruth_labels, marker='o')

        ari, ami, fmi, vm, ss = calculate_metrics(labels_gt, labels_gaussian, print_metrics_in_console=False,
                                                  ari_metric=ari_metric, ami_metric=ami_metric, fmi_metric=fmi_metric,
                                                  vm_metric=vm_metric, ss_metric=ss_metric)

        if avg_over_med:
            if ari:
                ari_avg += (ari / average)
            if ami:
                ami_avg += (ami / average)
            if fmi:
                fmi_avg += (fmi / average)
            if vm:
                vm_avg += (vm / average)
            if ss:
                ss_avg += (ss / average)
        else:
            if ari:
                metrics_values_list[0].append(ari)
            if ami:
                metrics_values_list[1].appeend(ami)
            if fmi:
                metrics_values_list[2].append(fmi)
            if vm:
                metrics_values_list[3].append(vm)
            if ss:
                metrics_values_list[4].append(ss)

        if write_intermediate_results:
            write_intermediate_metrics_excel(nr_of_dim=write_intermediate_results[0],
                                             sim=write_intermediate_results[1],
                                             nr_of_decimals=3, ari=ari, ami=ami, fmi=fmi, vm=vm, ss=ss,
                                             path=write_intermediate_results[2])

    if avg_over_med:
        return ari_avg, ami_avg, fmi_avg, vm_avg, ss_avg
    else:
        ari_median = statistics.median(metrics_values_list[0])
        ami_median = statistics.median(metrics_values_list[1])
        fmi_median = statistics.median(metrics_values_list[2])
        vm_median = statistics.median(metrics_values_list[3])
        ss_median = statistics.median(metrics_values_list[4])
        return ari_median, ami_median, fmi_median, vm_median, ss_median


def find_optimal_max_eps(folder_path='C:/Users/Cosmy/Desktop/Proiect Master/Rezultate simulari valori optime max_eps/'
                                     'sim',
                         sim_tuple=(1, 2, 3, 5, 13, 15, 34), dim_interval=(10, 21), nr_of_decimals=3):
    """
    Perform the OPTICS clustering with multiple values of the max_eps parameter for each dimension of each simulation,
    saving the results at the chosen path.

    :param folder_path: The path to the folder where the results will be saved
    :param sim_tuple: The tuple containing elements representing the number of the simulations to be clustered
    :param dim_interval: The tuple containing two elements, representing the start and the end of the interval of
    dimensions to be considered for each simulation
    :param nr_of_decimals: The number of decimals that the saved results should have
    :return:
    """
    for sim in sim_tuple:
        print(sim)
        for nr_of_dim in range(dim_interval[0], dim_interval[1]):
            print(nr_of_dim)
            path_to_save_folder = folder_path + str(sim) + '/' + str(nr_of_dim) + '/gt.png'
            f = open(folder_path + str(sim) + '/' + str(nr_of_dim) + '/metrics_max_eps.txt', "w")

            spikes_GT_2D, labels_GT_2D = get_dataset_simulation_pca_x_dimensions(sim_nr=sim, nr_of_dimensions_x=2)
            spikes, labels = get_dataset_simulation_pca_x_dimensions(sim_nr=sim, nr_of_dimensions_x=nr_of_dim)

            scatter_plot.plot('Ground Truth', spikes_GT_2D, labels_GT_2D, marker='o')
            plt.savefig(path_to_save_folder.strip())
            plt.close()

            max_ari, max_ami, max_fmi, max_vm, max_ss = 0, 0, 0, 0, 0
            ari_eps, ami_eps, fmi_eps, vm_eps, ss_eps = 0, 0, 0, 0, 0
            for max_eps in np.arange(0.5, 1.05, 0.05):
                optics = OPTICS(min_samples=35, max_eps=max_eps).fit(spikes)
                labels_optics = optics.labels_

                # print("\n", nr_of_dim, "D", max_eps, ":")
                ari, ami, fmi, vm, ss = calculate_metrics(labels_GT_2D, labels_optics, print_metrics_in_console=False,
                                                          ari_metric=True,
                                                          ami_metric=True,
                                                          fmi_metric=True, vm_metric=True, ss_metric=True)
                f.write("\n" + str(nr_of_dim) + "D " + str(max_eps) + " :")
                f.write('\nAdjusted Rand Index (ARI):' + str(round(ari, nr_of_decimals)) + '\n' +
                        'Adjusted Mutual Info Score (AMI):' + str(round(ami, nr_of_decimals)) + '\n' +
                        'Fowlkes Mallows Index (FMI):' + str(round(fmi, nr_of_decimals)) + '\n' +
                        'V-Measure (VM):' + str(round(vm, nr_of_decimals)) + '\n' +
                        'SS-Metric:' + str(round(ss, nr_of_decimals)) + '\n')

                if max_ari < ari:
                    max_ari = ari
                    ari_eps = max_eps
                if max_ami < ami:
                    max_ami = ami
                    ami_eps = max_eps
                if max_fmi < fmi:
                    max_fmi = fmi
                    fmi_eps = max_eps
                if max_vm < vm:
                    max_vm = vm
                    vm_eps = max_eps
                if max_ss < ss:
                    max_ss = ss
                    ss_eps = max_eps

                scatter_plot.plot(('OPTICS', nr_of_dim, 'D max_eps=', max_eps), spikes_GT_2D, labels_optics, marker='o')
                plt.savefig(folder_path + str(sim) + '/' + str(nr_of_dim) + '/' + str(max_eps) + '.png'.strip())
                plt.close()

            f.write(str(ari_eps) + str(ami_eps) + str(fmi_eps) + str(vm_eps) + str(ss_eps))
            f.close()


def full_clustering_process(path_optics_metrics='C:/Users/Cosmy/Desktop/Proiect Master/optics_metrics.xlsx',
                            path_gaussian_metrics='C:/Users/Cosmy/Desktop/Proiect Master/gaussian_metrics_100_'
                                                  'iterations_average.xlsx',
                            path_wir='C:/Users/Cosmy/Desktop/Proiect Master/intermediate_results_gaussian_clustering_'
                                     '100_iterations.xlsx',
                            average=100, avg_over_med=True):
    """
    Perform the full clustering process of multiple simulations, for multiple dimensions. This process involves an
    OPTICS clustering performed on the simulation and dimension of interest, after which one of the resulted clusters is
    selected (this cluster should be the main underclustered cluster for best results) and reclustered using gaussian
    clustering. The gaussian clustering of the chosen cluster is repeated multiple times and averages/median values
    of the resulted metrics are calculated

    :param path_optics_metrics: The path where the values of the metrics for the optics clustering process are saved
    :param path_gaussian_metrics: The path where the averages/median values of the metrics for the gaussian clustering
    process are saved
    :param path_wir: The path where the intermediate results (the results of each gaussian clustering) are saved
    :param average: The number of gaussian clusterings from which each average/median value is calculated
    :param avg_over_med: The parameter that decides whether the averages or the median values of the metrics for the
    gaussian clusterings will be calculated. The default value "True" selects the average, while the value "False"
    selects the median value
    :return:
    """
    # parameters [min_samples, max_eps, nr_components, cluster_to_apply_gaussian_reclustering_on], the row order is
    # the same as the order of the simulations in the list [1,2,3,5,13,14,34] and the order on each row represents
    # the parameters from 10D (first) to 20D (last)
    parameters = [
        [[35, 0.75, 8, 0], [35, 0.75, 8, 0], [35, 0.75, 8, 0], [35, 0.85, 8, 0], [35, 0.85, 8, 0], [35, 0.85, 8, 0],
         [35, 0.85, 8, 0], [35, 0.85, 8, 0], [35, 0.9, 8, 0], [35, 0.9, 8, 0], [35, 0.9, 8, 0]],
        [[35, 0.7, 13, 0], [35, 0.7, 11, 1], [35, 0.75, 12, 0], [35, 0.75, 11, 1], [35, 0.8, 13, 0], [35, 0.8, 12, 1],
         [35, 0.8, 11, 1], [35, 0.8, 10, 1], [35, 0.85, 10, 1], [35, 0.85, 10, 1], [35, 0.85, 10, 1]],
        [[35, 0.75, 6, 2], [35, 0.75, 6, 2], [35, 0.75, 6, 2], [35, 0.8, 6, 2], [35, 0.85, 8, 1], [35, 0.85, 8, 1],
         [35, 0.85, 8, 1], [35, 0.85, 6, 2], [35, 0.85, 6, 2], [35, 0.95, 8, 1], [35, 0.95, 8, 1]],
        [[35, 0.8, 10, 1], [35, 0.8, 9, 0], [35, 0.8, 10, 1], [35, 0.8, 10, 1], [35, 0.85, 10, 1], [35, 0.85, 10, 1],
         [35, 0.85, 10, 1], [35, 0.85, 10, 1], [35, 0.85, 10, 1], [35, 0.85, 10, 1], [35, 0.85, 10, 1]],
        [[35, 0.95, 9, 0], [35, 0.95, 9, 0], [35, 0.95, 6, 0], [35, 0.95, 6, 0], [35, 1, 6, 0], [35, 1, 6, 0],
         [35, 1, 6, 0], [35, 1, 6, 0], [35, 1.05, 6, 0], [35, 1.05, 6, 0], [35, 1.05, 6, 0]],
        [[35, 0.65, 2, 0], [35, 0.7, 2, 0], [35, 0.7, 2, 0], [35, 0.75, 2, 0], [35, 0.75, 2, 0], [35, 0.75, 2, 0],
         [35, 0.75, 2, 0], [35, 0.75, 2, 0], [35, 0.8, 2, 0], [35, 0.8, 2, 0], [35, 0.8, 2, 0]],
        [[35, 0.7, 2, 1], [35, 0.7, 2, 1], [35, 0.7, 2, 1], [35, 0.75, 2, 1], [35, 0.75, 2, 1], [35, 0.75, 2, 1],
         [35, 0.8, 2, 1], [35, 0.8, 2, 1], [35, 0.8, 2, 1], [35, 0.8, 2, 1], [35, 0.85, 2, 1]],
    ]
    parameters_row = -1
    for sim in [1, 2, 3, 5, 13, 14, 34]:
        # print("sim ", sim)
        parameters_row += 1
        parameters_column = -1
        best_ss = 0
        best_nr_of_dim: list = []
        for nr_of_dim in [10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]:
            # print("nr of dim ", nr_of_dim)
            parameters_column += 1
            labels_of_underclustered_cluster = parameters[parameters_row][parameters_column][3]
            spikes_GT_2D, labels_GT_2D = get_dataset_simulation_pca_x_dimensions(sim_nr=sim, nr_of_dimensions_x=2)
            spikes, labels = get_dataset_simulation_pca_x_dimensions(sim_nr=sim, nr_of_dimensions_x=nr_of_dim)
            optics = OPTICS(min_samples=parameters[parameters_row][parameters_column][0],
                            max_eps=parameters[parameters_row][parameters_column][1]).fit(spikes)
            labels_optics = optics.labels_

            # print("\n", nr_of_dim, "D:")
            ari, ami, fmi, vm, ss = calculate_metrics(labels_GT_2D, labels_optics, print_metrics_in_console=False,
                                                      ari_metric=True, ami_metric=True,
                                                      fmi_metric=True, vm_metric=True, ss_metric=True)

            write_metrics_excel(nr_of_dim, sim, 3, ari, ami, fmi, vm, ss, path_optics_metrics)

            # secondary clustering, GAUSSIAN
            spikes79, labels = ds.get_dataset_simulation(sim)
            current_cluster = spikes79[labels_optics == labels_of_underclustered_cluster]

            current_cluster_2d = spikes_GT_2D[labels_optics == labels_of_underclustered_cluster]
            current_groundtruth_labels = labels_GT_2D[labels_optics == labels_of_underclustered_cluster]

            # scatter_plot.plot('GT cluster', current_cluster_2d, current_groundtruth_labels, marker='o')

            pca_xd = pca.PCA(n_components=nr_of_dim)
            spikes79_pca = pca_xd.fit_transform(current_cluster)

            pca_xd = pca.PCA(n_components=2)
            spikes_pca_2d = pca_xd.fit_transform(current_cluster)

            ari_avg, ami_avg, fmi_avg, vm_avg, ss_avg = average_metrics(spikes79_pca, current_groundtruth_labels,
                                                                        spikes_pca_2d, average=average,
                                                                        avg_over_med=avg_over_med,
                                                                        nr_of_subclusters=
                                                                        parameters[parameters_row][parameters_column][
                                                                            2],
                                                                        write_intermediate_results=[nr_of_dim, sim,
                                                                                                    path_wir],
                                                                        ari_metric=True, ami_metric=True,
                                                                        fmi_metric=True, vm_metric=True, ss_metric=True)

            write_metrics_excel(nr_of_dim, sim, 3, ari_avg, ami_avg, fmi_avg, vm_avg, ss_avg, path_gaussian_metrics)

            if ss_avg > best_ss:
                best_ss = ss_avg
                best_nr_of_dim.clear()
                best_nr_of_dim.append(nr_of_dim)
            elif ss_avg == best_ss:
                best_nr_of_dim.append(nr_of_dim)

        # print("For sim", sim, "the best dimensions are", best_nr_of_dim, "with ss score", best_ss)
    plt.show()


if __name__ == "__main__":
    ari_values = np.array(
        [0.716, 0.716, 0.584, 0.716, 0.717, 0.584, 0.584, 0.716, 0.584, 0.716, 0.664, 0.584, 0.584, 0.584, 0.717, 0.584,
         0.716, 0.584, 0.689, 0.584, 0.584, 0.584, 0.689, 0.689, 0.689, 0.584, 0.614, 0.716, 0.664, 0.689, 0.584, 0.716,
         0.689, 0.661, 0.632, 0.584, 0.664, 0.716, 0.716, 0.584, 0.584, 0.455, 0.584, 0.584, 0.689, 0.718, 0.716, 0.716,
         0.716, 0.584, 0.584, 0.584, 0.716, 0.716, 0.584, 0.716, 0.664, 0.716, 0.584, 0.584, 0.584, 0.584, 0.668, 0.584,
         0.716, 0.584, 0.717, 0.584, 0.584, 0.689, 0.584, 0.716, 0.451, 0.584, 0.716, 0.584, 0.614, 0.584, 0.717, 0.584,
         0.584, 0.716, 0.716, 0.716, 0.615, 0.716, 0.584, 0.584, 0.584, 0.664, 0.584, 0.584, 0.584, 0.664, 0.584, 0.716,
         0.689, 0.584, 0.716, 0.584])

    ami_values = np.array(
        [0.838, 0.838, 0.781, 0.838, 0.839, 0.781, 0.781, 0.838, 0.781, 0.838, 0.815, 0.781, 0.781, 0.781, 0.839, 0.781,
         0.838, 0.781, 0.834, 0.781, 0.781, 0.781, 0.834, 0.834, 0.834, 0.781, 0.783, 0.838, 0.815, 0.834, 0.781, 0.838,
         0.834, 0.813, 0.794, 0.781, 0.815, 0.838, 0.838, 0.781, 0.781, 0.726, 0.781, 0.781, 0.834, 0.839, 0.838, 0.838,
         0.838, 0.781, 0.781, 0.781, 0.838, 0.838, 0.781, 0.838, 0.816, 0.838, 0.781, 0.781, 0.781, 0.781, 0.819, 0.781,
         0.838, 0.781, 0.838, 0.781, 0.781, 0.835, 0.781, 0.838, 0.724, 0.781, 0.838, 0.781, 0.782, 0.781, 0.838, 0.781,
         0.781, 0.838, 0.838, 0.838, 0.784, 0.838, 0.781, 0.781, 0.781, 0.815, 0.781, 0.781, 0.781, 0.815, 0.781, 0.838,
         0.835, 0.781, 0.838, 0.781])

    fmi_values = np.array(
        [0.78, 0.78, 0.669, 0.78, 0.78, 0.669, 0.669, 0.78, 0.669, 0.78, 0.732, 0.669, 0.669, 0.669, 0.78, 0.669, 0.78,
         0.669, 0.754, 0.669, 0.669, 0.669, 0.754, 0.754, 0.754, 0.669, 0.691, 0.78, 0.732, 0.754, 0.669, 0.78, 0.754,
         0.73, 0.706, 0.669, 0.732, 0.78, 0.78, 0.669, 0.669, 0.559, 0.669, 0.669, 0.754, 0.781, 0.78, 0.78, 0.78,
         0.669, 0.669, 0.669, 0.78, 0.78, 0.669, 0.78, 0.733, 0.78, 0.669, 0.669, 0.669, 0.669, 0.736, 0.669, 0.78,
         0.669, 0.78, 0.669, 0.669, 0.754, 0.669, 0.78, 0.556, 0.669, 0.78, 0.669, 0.691, 0.669, 0.78, 0.669, 0.669,
         0.78, 0.78, 0.78, 0.691, 0.78, 0.669, 0.669, 0.669, 0.732, 0.669, 0.669, 0.669, 0.732, 0.669, 0.78, 0.754,
         0.669, 0.78, 0.669])

    vm_values = np.array(
        [0.884, 0.884, 0.829, 0.883, 0.884, 0.829, 0.829, 0.883, 0.829, 0.884, 0.848, 0.829, 0.829, 0.829, 0.884, 0.829,
         0.883, 0.829, 0.868, 0.829, 0.829, 0.829, 0.868, 0.868, 0.868, 0.829, 0.815, 0.884, 0.848, 0.867, 0.829, 0.884,
         0.868, 0.847, 0.82, 0.829, 0.848, 0.884, 0.884, 0.829, 0.829, 0.755, 0.829, 0.829, 0.868, 0.885, 0.883, 0.884,
         0.883, 0.829, 0.829, 0.829, 0.884, 0.884, 0.829, 0.884, 0.849, 0.884, 0.829, 0.829, 0.829, 0.829, 0.85, 0.829,
         0.884, 0.829, 0.884, 0.829, 0.829, 0.868, 0.829, 0.884, 0.753, 0.829, 0.884, 0.829, 0.814, 0.829, 0.884, 0.829,
         0.829, 0.884, 0.884, 0.884, 0.816, 0.884, 0.829, 0.829, 0.829, 0.848, 0.829, 0.829, 0.829, 0.848, 0.829, 0.883,
         0.868, 0.829, 0.884, 0.829
         ])

    ss_values = np.array(
        [0.62, 0.62, 0.542, 0.62, 0.621, 0.542, 0.542, 0.62, 0.542, 0.62, 0.543, 0.542, 0.542, 0.542, 0.621, 0.542,
         0.62, 0.542, 0.611, 0.542, 0.542, 0.542, 0.611, 0.611, 0.611, 0.542, 0.533, 0.62, 0.543, 0.611, 0.542, 0.62,
         0.611, 0.543, 0.533, 0.542, 0.543, 0.62, 0.62, 0.542, 0.542, 0.452, 0.542, 0.542, 0.611, 0.621, 0.62, 0.62,
         0.62, 0.542, 0.542, 0.542, 0.62, 0.62, 0.542, 0.62, 0.542, 0.62, 0.542, 0.542, 0.542, 0.542, 0.544, 0.542,
         0.62, 0.542, 0.62, 0.542, 0.542, 0.611, 0.542, 0.62, 0.452, 0.542, 0.62, 0.542, 0.532, 0.542, 0.62, 0.542,
         0.542, 0.62, 0.62, 0.62, 0.535, 0.62, 0.542, 0.542, 0.542, 0.542, 0.542, 0.542, 0.542, 0.543, 0.542, 0.62,
         0.611, 0.542, 0.62, 0.542
         ])

    # Creating histogram
    fig, ax = plt.subplots(figsize=(10, 7))
    ax.hist(ss_values, bins=[0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1])

    fig2, ay = plt.subplots(figsize=(10, 7))
    ay.hist(vm_values, bins=[0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1])

    # Show plot
    plt.show()
